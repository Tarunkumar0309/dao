from openpyxl import load_workbook, Workbook

def readDataFromExcel(filename):
    wb = load_workbook(filename)
    sheet = wb.active
    data = list(sheet.iter_rows(values_only=True))
    return data